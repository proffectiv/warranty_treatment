#!/usr/bin/env python3
"""
Analyze Excel structure to find empty rows and understand max_row behavior
"""

import os
import sys
from dotenv import load_dotenv

# Import our modules
from update_excel_dropbox import get_dropbox_access_token, download_excel_from_dropbox
from openpyxl import load_workbook

# Import logging filter from root directory
sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..'))
from log_filter import setup_secure_logging

load_dotenv()

# Set up secure logging
logger = setup_secure_logging('excel_analysis')

def analyze_excel_structure(brand="Conway"):
    """
    Analyze the Excel structure to understand row usage and empty rows
    """
    try:
        # Get Dropbox credentials
        access_token = get_dropbox_access_token()
        folder_path = os.getenv('DROPBOX_FOLDER_PATH')
        file_path = f"{folder_path}/GARANTIAS_PROFFECTIV.xlsx"
        
        logger.info(f"Downloading Excel file from: {file_path}")
        
        # Download existing Excel file
        excel_file = download_excel_from_dropbox(access_token, file_path)
        
        # Load workbook with openpyxl
        workbook = load_workbook(excel_file, data_only=True)
        
        if brand not in workbook.sheetnames:
            logger.error(f"Sheet '{brand}' not found in Excel file")
            return False
        
        # Get the specific brand worksheet
        worksheet = workbook[brand]
        
        logger.info(f"Analyzing {brand} sheet:")
        logger.info(f"Max row (openpyxl): {worksheet.max_row}")
        logger.info(f"Max column (openpyxl): {worksheet.max_column}")
        
        # Count rows with data vs empty rows
        rows_with_data = 0
        empty_rows = 0
        last_data_row = 1
        
        # Sample every 10th row to find patterns
        sample_rows = list(range(1, min(worksheet.max_row + 1, 100))) + \
                      list(range(100, worksheet.max_row + 1, 10))
        
        logger.info("Sampling rows to find data patterns:")
        
        for row in sample_rows:
            # Check if row has any data (check first 5 columns)
            has_data = False
            for col in range(1, 6):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value is not None and str(cell_value).strip():
                    has_data = True
                    break
            
            if has_data:
                rows_with_data += 1
                last_data_row = row
                if row <= 20 or row % 100 == 0 or row > worksheet.max_row - 10:
                    logger.info(f"Row {row}: HAS DATA")
            else:
                empty_rows += 1
                if row <= 20 or row % 100 == 0 or row > worksheet.max_row - 10:
                    logger.info(f"Row {row}: EMPTY")
        
        logger.info(f"Summary of sampled rows:")
        logger.info(f"- Rows with data: {rows_with_data}")
        logger.info(f"- Empty rows: {empty_rows}")
        logger.info(f"- Last row with data: {last_data_row}")
        
        # Check if there are formatting or hidden elements causing high max_row
        logger.info("Checking for potential causes of high max_row:")
        
        # Check merged cells
        merged_ranges = list(worksheet.merged_cells.ranges)
        if merged_ranges:
            logger.info(f"Found {len(merged_ranges)} merged cell ranges")
            for i, merged in enumerate(merged_ranges[:5]):  # Show first 5
                logger.info(f" - Merged range {i+1}: {merged}")
        
        # Check if there are any cells with formatting but no data in high rows
        test_rows = [worksheet.max_row - i for i in range(5)]
        logger.info(f"Checking formatting in last 5 rows: {test_rows}")
        
        for row in test_rows:
            row_info = []
            for col in range(1, 6):
                cell = worksheet.cell(row=row, column=col)
                has_value = cell.value is not None
                has_format = (cell.font.name != 'Calibri' or 
                             cell.fill.fgColor.rgb != '00000000' or
                             cell.border.top.style is not None)
                row_info.append(f"Col{col}:{'V' if has_value else 'E'}{'F' if has_format else ''}")
            
            logger.info(f"Row {row}: {' '.join(row_info)}")
        
        return True
        
    except Exception as e:
        logger.error(f"Error analyzing Excel structure: {str(e)}")
        return False

if __name__ == "__main__":
    brand = sys.argv[1] if len(sys.argv) > 1 else "Conway"
    analyze_excel_structure(brand)