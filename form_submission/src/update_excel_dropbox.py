import os
import json
import requests
import pandas as pd
import sys
from datetime import datetime, timedelta
from io import BytesIO
from dotenv import load_dotenv
# SequenceMatcher removed - no longer needed
from openpyxl import load_workbook
from openpyxl.styles import Font

# Import logging filter from root directory
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..'))
from log_filter import setup_secure_logging
from warranty_form_data import WarrantyFormData

load_dotenv()

# Set up secure logging
logger = setup_secure_logging('excel_dropbox')

def get_dropbox_access_token():
    """Get access token from refresh token"""
    refresh_token = os.getenv('DROPBOX_REFRESH_TOKEN')
    app_key = os.getenv('DROPBOX_APP_KEY')
    app_secret = os.getenv('DROPBOX_APP_SECRET')
    
    url = 'https://api.dropbox.com/oauth2/token'
    data = {
        'grant_type': 'refresh_token',
        'refresh_token': refresh_token,
        'client_id': app_key,
        'client_secret': app_secret
    }
    
    response = requests.post(url, data=data)
    if response.status_code == 200:
        return response.json()['access_token']
    else:
        raise Exception(f"Failed to get access token: {response.text}")

def download_excel_from_dropbox(access_token, file_path):
    """Download Excel file from Dropbox"""
    url = 'https://content.dropboxapi.com/2/files/download'
    headers = {
        'Authorization': f'Bearer {access_token}',
        'Dropbox-API-Arg': json.dumps({'path': file_path})
    }
    
    response = requests.post(url, headers=headers)
    if response.status_code == 200:
        return BytesIO(response.content)
    else:
        raise Exception(f"Failed to download file: {response.text}")

def upload_excel_to_dropbox(access_token, file_path, excel_data):
    """Upload Excel file to Dropbox"""
    url = 'https://content.dropboxapi.com/2/files/upload'
    headers = {
        'Authorization': f'Bearer {access_token}',
        'Dropbox-API-Arg': json.dumps({
            'path': file_path,
            'mode': 'overwrite',
            'autorename': False
        }),
        'Content-Type': 'application/octet-stream'
    }
    
    response = requests.post(url, headers=headers, data=excel_data)
    if response.status_code == 200:
        return response.json()
    else:
        raise Exception(f"Failed to upload file: {response.text}")

# Remove the old field parsing function since we now use WarrantyFormData

# Duplicate detection functions removed as requested

# Remove the old prepare_row_data function since we now use WarrantyFormData.to_excel_row()

def find_column_index(worksheet, column_name):
    """Find the column index for a given column name in the header row"""
    for col_idx, cell in enumerate(worksheet[1], 1):
        if cell.value == column_name:
            return col_idx
    return None

# Removed find_first_empty_row_within_validation and extend_data_validation_range functions
# since we now use worksheet.insert_rows() which automatically handles data validation ranges

def update_excel_file(form_data: WarrantyFormData):
    """Main function to update Excel file in Dropbox using WarrantyFormData object with openpyxl to preserve formatting"""
    try:
        brand = form_data.brand
        
        if not brand or brand == 'No especificado':
            raise Exception("Brand not found in form data")
        
        logger.info(f"Processing Excel update for brand: {brand}")
        logger.info(f"Form data ticket ID: {form_data.ticket_id}")
        
        # Get Dropbox credentials
        access_token = get_dropbox_access_token()
        folder_path = os.getenv('DROPBOX_FOLDER_PATH')
        file_path = f"{folder_path}/GARANTIAS_PROFFECTIV.xlsx"
        
        logger.info(f"Downloading Excel file from: {file_path}")
        
        # Download existing Excel file
        excel_file = download_excel_from_dropbox(access_token, file_path)
        
        # Load workbook with openpyxl to preserve formatting and data validation
        workbook = load_workbook(excel_file, data_only=False)
        
        logger.info(f"Available sheets in workbook: {workbook.sheetnames}")
        
        # Check if brand sheet exists
        if brand not in workbook.sheetnames:
            raise Exception(f"Sheet '{brand}' not found in Excel file")
        
        # Get the specific brand worksheet
        worksheet = workbook[brand]
        
        # Get column headers from first row to map data correctly
        headers = {}
        for col_idx, cell in enumerate(worksheet[1], 1):
            if cell.value:
                headers[cell.value] = col_idx
        
        logger.info(f"Excel headers found: {list(headers.keys())}")
        
        # Find all rows with ticket ID data to understand the data structure
        ticket_id_col = headers.get('Ticket ID', 1)
        
        # NEW APPROACH: Always insert at row 2 (after headers) and shift existing data down
        insert_row = 2
        logger.info(f"Will insert new row at position {insert_row} (after headers)")
        
        # Find the last row with actual data to know how many rows to shift
        data_rows = []
        scan_range = min(worksheet.max_row + 1, 2000)  # Limit scan to prevent performance issues
        
        for row in range(2, scan_range):
            ticket_id = worksheet.cell(row=row, column=ticket_id_col).value
            if ticket_id and str(ticket_id).strip():
                data_rows.append(row)
        
        if data_rows:
            actual_last_row = max(data_rows)
            logger.info(f"Found existing data up to row {actual_last_row}. Will shift existing data down by 1 row.")
            
            # Insert a new row at position 2, which automatically shifts existing data down
            worksheet.insert_rows(insert_row, 1)
            logger.info(f"Inserted new row at position {insert_row}")
        else:
            logger.info("No existing data found, will add first data row at position 2")
        
        # The row to write to is always row 2 after insertion
        next_row = insert_row
        
        # Prepare new row data using form_data
        new_row_data = form_data.to_excel_row(brand)
        
        logger.info(f"New row data keys: {list(new_row_data.keys())}")
        logger.info(f"New row data values: {list(new_row_data.values())}")
        
        # Write new row data to worksheet
        cells_written = 0
        estado_col_idx = None
        for column_name, value in new_row_data.items():
            if column_name in headers:
                col_idx = headers[column_name]
                cell = worksheet.cell(row=next_row, column=col_idx)
                old_value = cell.value
                
                # Handle hyperlink formatting for URL fields
                if isinstance(value, dict) and value.get('type') == 'hyperlink':
                    cell.value = value['text']
                    cell.hyperlink = value['url']
                    cell.font = Font(color="0000FF", underline="single")  # Blue underlined text
                    logger.info(f"Writing hyperlink to cell {cell.coordinate}: '{value['text']}' -> '{value['url']}'")
                else:
                    cell.value = value
                    logger.info(f"Writing to cell {cell.coordinate} (row {next_row}, col {col_idx}): '{column_name}' = '{value}' (was: '{old_value}')")
                
                cells_written += 1
                
                # Remember Estado column for validation extension
                if column_name == 'Estado':
                    estado_col_idx = col_idx
                
            else:
                logger.warning(f"Column '{column_name}' not found in Excel headers")
        
        # After writing all data, extend data validation for Estado column if needed
        # Since we inserted a row, the data validation ranges are automatically extended by openpyxl
        # but we may need to verify and adjust if necessary
        if estado_col_idx:
            logger.info(f"Data validation should be automatically extended due to row insertion at {next_row}")
            # The extend_data_validation_range function is no longer needed since insert_rows handles this
            # extend_data_validation_range(worksheet, estado_col_idx, next_row)
        
        logger.info(f"Total cells written: {cells_written}")
        
        # Verify the data was written by checking a few key cells
        verification_cells = ['Ticket ID', 'Estado', 'Empresa']
        for col_name in verification_cells:
            if col_name in headers:
                col_idx = headers[col_name]
                cell_value = worksheet.cell(row=next_row, column=col_idx).value
                logger.info(f"Verification - {col_name}: '{cell_value}'")
        
        # Check new max row after writing
        new_max_row = worksheet.max_row
        logger.info(f"New max row after writing: {new_max_row} (should be {next_row})")
        
        # Save workbook to BytesIO
        logger.info("Saving workbook to BytesIO...")
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Get size of saved data
        saved_size = len(output.getvalue())
        logger.info(f"Saved Excel file size: {saved_size} bytes")
        
        if saved_size == 0:
            raise Exception("Saved Excel file is empty!")
        
        # Upload updated file back to Dropbox
        logger.info("Uploading file to Dropbox...")
        upload_result = upload_excel_to_dropbox(access_token, file_path, output.getvalue())
        
        logger.info(f"Excel file updated successfully. New row added to {brand} sheet.")
        logger.info(f"File uploaded to Dropbox: {upload_result.get('path_display', file_path)}")
        
        return True
        
    except Exception as e:
        logger.error(f"Error updating Excel file: {str(e)}")
        return False

if __name__ == "__main__":
    # Test with sample data
    import sys
    if len(sys.argv) > 1:
        with open(sys.argv[1], 'r') as f:
            webhook_data = json.load(f)
        form_data = WarrantyFormData(webhook_data, "test-ticket-123")
        update_excel_file(form_data)