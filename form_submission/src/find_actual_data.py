#!/usr/bin/env python3
"""
Find where actual data starts and ends in the Excel file
"""

import os
import sys
from dotenv import load_dotenv

# Import our modules
from update_excel_dropbox import get_dropbox_access_token, download_excel_from_dropbox
from openpyxl import load_workbook

# Import logging filter from root directory
sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..'))
from log_filter import setup_secure_logging

load_dotenv()

# Set up secure logging
logger = setup_secure_logging('find_data')

def find_actual_data(brand="Conway"):
    """
    Find where actual warranty data starts and ends
    """
    try:
        # Get Dropbox credentials
        access_token = get_dropbox_access_token()
        folder_path = os.getenv('DROPBOX_FOLDER_PATH')
        file_path = f"{folder_path}/GARANTIAS_PROFFECTIV.xlsx"
        
        logger.info(f"Downloading Excel file from: {file_path}")
        
        # Download existing Excel file
        excel_file = download_excel_from_dropbox(access_token, file_path)
        
        # Load workbook with openpyxl
        workbook = load_workbook(excel_file, data_only=True)
        
        if brand not in workbook.sheetnames:
            logger.error(f"Sheet '{brand}' not found in Excel file")
            return False
        
        # Get the specific brand worksheet
        worksheet = workbook[brand]
        
        logger.info(f"Analyzing {brand} sheet (max_row: {worksheet.max_row}):")
        
        # Find first and last row with ticket ID data
        ticket_id_rows = []
        
        # Check every row for Ticket ID in column A
        for row in range(1, worksheet.max_row + 1):
            ticket_id = worksheet.cell(row=row, column=1).value
            if ticket_id and str(ticket_id).strip() and ticket_id != 'Ticket ID':
                ticket_id_rows.append(row)
        
        if not ticket_id_rows:
            logger.error("No ticket ID data found!")
            return False
        
        first_data_row = min(ticket_id_rows)
        last_data_row = max(ticket_id_rows)
        total_data_rows = len(ticket_id_rows)
        
        logger.info(f"Data analysis:")
        logger.info(f"- First data row: {first_data_row}")
        logger.info(f"- Last data row: {last_data_row}")
        logger.info(f"- Total rows with ticket IDs: {total_data_rows}")
        logger.info(f"- Expected next row: {last_data_row + 1}")
        
        # Show first few and last few data rows
        logger.info(f"First 3 data rows:")
        for i, row in enumerate(ticket_id_rows[:3]):
            ticket_id = worksheet.cell(row=row, column=1).value
            estado = worksheet.cell(row=row, column=2).value
            empresa = worksheet.cell(row=row, column=4).value
            logger.info(f"  Row {row}: {ticket_id} | {estado} | {empresa}")
        
        logger.info(f"Last 3 data rows:")
        for i, row in enumerate(ticket_id_rows[-3:]):
            ticket_id = worksheet.cell(row=row, column=1).value
            estado = worksheet.cell(row=row, column=2).value
            empresa = worksheet.cell(row=row, column=4).value
            logger.info(f"  Row {row}: {ticket_id} | {estado} | {empresa}")
        
        # Check for gaps in the data
        gaps = []
        for i in range(len(ticket_id_rows) - 1):
            current_row = ticket_id_rows[i]
            next_row = ticket_id_rows[i + 1]
            if next_row - current_row > 1:
                gaps.append((current_row, next_row))
        
        if gaps:
            logger.info(f"Found {len(gaps)} gaps in data:")
            for start, end in gaps[:5]:  # Show first 5 gaps
                logger.info(f"  Gap: rows {start} to {end} ({end - start - 1} empty rows)")
        else:
            logger.info("No gaps found - data is continuous")
        
        return True
        
    except Exception as e:
        logger.error(f"Error finding actual data: {str(e)}")
        return False

if __name__ == "__main__":
    brand = sys.argv[1] if len(sys.argv) > 1 else "Conway"
    find_actual_data(brand)