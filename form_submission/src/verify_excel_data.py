#!/usr/bin/env python3
"""
Verification script to check if data is actually in the Excel file
"""

import os
import json
import sys
from dotenv import load_dotenv

# Import our modules
from update_excel_dropbox import get_dropbox_access_token, download_excel_from_dropbox
from openpyxl import load_workbook

# Import logging filter from root directory
sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..'))
from log_filter import setup_secure_logging

load_dotenv()

# Set up secure logging
logger = setup_secure_logging('excel_verification')

def verify_excel_data(brand="Conway", last_n_rows=10):
    """
    Verify that data exists in the Excel file by downloading and checking it
    
    Args:
        brand: Brand sheet to check
        last_n_rows: Number of last rows to display
    """
    try:
        # Get Dropbox credentials
        access_token = get_dropbox_access_token()
        folder_path = os.getenv('DROPBOX_FOLDER_PATH')
        file_path = f"{folder_path}/GARANTIAS_PROFFECTIV.xlsx"
        
        logger.info(f"Downloading Excel file from: {file_path}")
        
        # Download existing Excel file
        excel_file = download_excel_from_dropbox(access_token, file_path)
        
        # Load workbook with openpyxl
        workbook = load_workbook(excel_file, data_only=True)  # data_only=True to get calculated values
        
        logger.info(f"Available sheets: {workbook.sheetnames}")
        
        if brand not in workbook.sheetnames:
            logger.error(f"Sheet '{brand}' not found in Excel file")
            return False
        
        # Get the specific brand worksheet
        worksheet = workbook[brand]
        
        # Get headers
        headers = []
        for col_idx, cell in enumerate(worksheet[1], 1):
            if cell.value:
                headers.append(cell.value)
        
        logger.info(f"Headers in {brand} sheet: {headers}")
        logger.info(f"Total max row: {worksheet.max_row}")
        
        # Find actual last row with data
        actual_last_row = 1
        for row in range(2, worksheet.max_row + 1):
            ticket_id_cell = worksheet.cell(row=row, column=1)  # Assuming Ticket ID is in column A
            if ticket_id_cell.value is not None:
                actual_last_row = row
        
        logger.info(f"Actual last row with data: {actual_last_row}")
        
        # Show the last few rows
        start_row = max(2, actual_last_row - last_n_rows + 1)
        logger.info(f"Showing rows {start_row} to {actual_last_row}:")
        
        for row in range(start_row, actual_last_row + 1):
            row_data = []
            for col in range(1, len(headers) + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            
            # Show first few columns to verify data exists
            ticket_id = row_data[0] if len(row_data) > 0 else ""
            estado = row_data[1] if len(row_data) > 1 else ""
            empresa = row_data[3] if len(row_data) > 3 else ""
            
            logger.info(f"Row {row}: Ticket ID='{ticket_id}', Estado='{estado}', Empresa='{empresa[:20]}...'")
        
        logger.info(f"✅ Verification complete. Found {actual_last_row - 1} data rows in {brand} sheet.")
        return True
        
    except Exception as e:
        logger.error(f"Error verifying Excel data: {str(e)}")
        return False

if __name__ == "__main__":
    # Allow specifying brand and number of rows to show
    brand = sys.argv[1] if len(sys.argv) > 1 else "Conway"
    last_n_rows = int(sys.argv[2]) if len(sys.argv) > 2 else 5
    
    verify_excel_data(brand, last_n_rows)